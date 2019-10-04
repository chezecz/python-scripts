import openpyxl

def get_items():
	wb = openpyxl.load_workbook('../Book1.xlsx')
	ws = wb['Sheet1']

	item_list = []

	for i in range(2, ws.max_row):
		item_list.append("'" + str(ws.cell(column = 1, row = i).value) + "'")

	return item_list	

def create_query(item_list):
	query = f'''update unitcd4
		set description = 'false'
		where unit4 IN ({','.join(item_list)})'''

	return query

if __name__ == '__main__':
	print(create_query(get_items()))
