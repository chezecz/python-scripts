import openpyxl

wb = openpyxl.load_workbook('../Book1.xlsx')

ws = wb['MiscellaneousIssuesReasonCodesE']

items = []
description = []

def construct_sql_query(items, description):
    for item in items:
        sql_string = f"update reason\nset description = {item[1]}\nwhere reason_code = {item[0]} and reason_class = 'MISC RCPT'"
        print(sql_string)

for i in range(2, ws.max_row):
    items.append(("'" + str(ws.cell(column = 1, row = i).value) + "'", "'" + str(ws.cell(column = 2, row = i).value) + "'"))

construct_sql_query(items, description)