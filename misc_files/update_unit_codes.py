import openpyxl
import re

wb = openpyxl.load_workbook('../../Stock.xlsx')

ws = wb['ItemStockroomLocationsExport31']

pattern = ""

def construct_query(warehouse, location, item, unit_code_2, lbr_code_2, fixed_unit_code_2, var_unit_code_2, out_unit_code_2):
    query = f"""update itemloc
                set inv_acct_unit2 = '{unit_code_2}',
                lbr_acct_unit2 = '{lbr_code_2}',
                fovhd_acct_unit2 = '{fixed_unit_code_2}',
                vovhd_acct_unit2 = '{var_unit_code_2}',
                out_acct_unit2 = '{out_unit_code_2}'
                where item = '{item}'
                and loc = '{location}'
                and whse = '{warehouse}'"""
    print(query)

for i in range(2, ws.max_row):
    warehouse = ws.cell(column = 1, row = i).value
    location = ws.cell(column = 2, row = i).value
    item = ws.cell(column = 3, row = i).value
    unit_code_2 = ws.cell(column = 6, row = i).value
    lbr_code_2 = ws.cell(column = 12, row = i).value
    fixed_unit_code_2 = ws.cell(column = 18, row = i).value
    var_unit_code_2 = ws.cell(column = 24, row = i).value
    out_unit_code_2 = ws.cell(column = 30, row = i).value

    # if (re.search(r"\B'\w+'\B", item)):
        # item = re.sub(r"\B'\w+'\B", "'"+str(item)+"'", item)

    if (unit_code_2 != ''):
        construct_query(warehouse, location, item, unit_code_2, lbr_code_2, fixed_unit_code_2, var_unit_code_2, out_unit_code_2)