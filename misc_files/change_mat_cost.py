import openpyxl

wb = openpyxl.load_workbook('../Book1.xlsx')

ws = wb['Sheet1']

def construct_sql_query(item, cost):
    sql_string = f"update item\nset cur_mat_cost = '{cost}'\nwhere item = '{item}'"
    print(sql_string)

def construct_exec_query(item):
    sql_string = f"exec Rpt_RollCurrentCosttoStandardCostSp @Post = 1 ,@FromProductCode = NULL,@ToProductCode = NULL,@FromItem = '{item}',@ToItem = '{item}',@Source = 'PMT',@ABC = 'ABC',@CostMethod = 'ALFSC',@MatlType = 'MTFO',@DisplayHeader = '1',@UserID = '2',@BGSessionId = NULL"
    print(sql_string)

for i in range(2, ws.max_row):
    #construct_sql_query(ws.cell(column = 1, row = i).value, round(ws.cell(column = 7, row = i).value, 4))
    construct_exec_query(ws.cell(column = 1, row = i).value)