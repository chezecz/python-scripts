import openpyxl

wb = openpyxl.load_workbook('../Book1.xlsx')

ws = wb['Sheet1']

def construct_sql_query(item, cost):
    sql_string = f"update reason\nset description = 'false'\nwhere item NOT IN ({','.join(items)})"
    print(sql_string)

for i in range(2, ws.max_row):
    construct_exec_query(ws.cell(column = 1, row = i).value)