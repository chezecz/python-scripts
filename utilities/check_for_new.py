import openpyxl

from datetime import datetime, timedelta

from excel.add_lines import from_excel_ordinal

from utilities.get_dates import check_week

def check_new(worksheet):
    worksheet_max_row = worksheet.max_row
    day_today = datetime.today()- timedelta(days = 1)
    day_today = day_today.strftime('%d%m%Y')
    if worksheet.cell(row = worksheet_max_row, column = 13).value == None:
        return False
    if from_excel_ordinal(worksheet.cell(row = worksheet_max_row, column = 13).value) == day_today or check_week():
        return True
    else:
        return False