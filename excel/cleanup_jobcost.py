import openpyxl
import re

from openpyxl.utils.cell import coordinate_from_string, column_index_from_string

def cleanup_file(worksheet):
    worksheet.delete_rows(1, 5)
    worksheet = swap_columns(worksheet)
    worksheet = delete_unnecessary_columns(worksheet)
    worksheet_max_row = worksheet.max_row  
    worksheet = remove_rows(worksheet, worksheet_max_row)
    return worksheet

def swap_columns(worksheet):
    worksheet.move_range("AK1:AM100", cols = 1)
    
    for src, dst in zip(worksheet['AN:AN'], worksheet['AK:AK']):
        dst.value = src.value
        
    return worksheet
    
def delete_unnecessary_columns(worksheet):

    # list of unnecessary columns
    
    columns_to_delete = []
    columns_to_delete.extend((
    column_index_from_string('B'),
    column_index_from_string('F'),
    column_index_from_string('G'),
    column_index_from_string('I'),
    column_index_from_string('K'),
    column_index_from_string('S'),
    column_index_from_string('AA'),
    column_index_from_string('AH'),
    column_index_from_string('AN')
    ))
    
    # delete unnecessary columns

    for element in reversed(columns_to_delete):
        worksheet.delete_cols(element, 1)
    
    return worksheet
    
def remove_rows(worksheet, worksheet_max_row):
    for i in range(worksheet_max_row + 1, 0, -1):
        if re.match("^\d{1,2}\/\d{1,2}\/\d{4}$", str(worksheet.cell(row = i, column = 1).value)):
            worksheet.cell(row = i, column = 1).value = ""
            continue
        if re.match("^[MCDJ]", str(worksheet.cell(row = i, column = 4).value)) == None or re.match("^TR", str(worksheet.cell(row = i, column = 1).value)):
            worksheet.delete_rows(i, 1)
        elif worksheet.cell(row = i, column = 12).value == 0:
            worksheet.delete_rows(i, 1)
    
    k = 0
    i = 1
    while worksheet.cell(row = i, column = 1).value == "":
        k += 1
        i += 1
        
    worksheet.delete_rows(1, k)

    return worksheet        