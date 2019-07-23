import openpyxl

from datetime import datetime, timedelta

from excel.sum_excel import get_column_sum, get_total_sum

def add_blank_lines(worksheet, max_row):
    worksheet_max_row = worksheet.max_row
    worksheet_max_column = worksheet.max_column
    
    for i in range(2, worksheet_max_row+1):
        if worksheet.cell(row = i+1, column = 14).value and worksheet.cell(row = i, column = 14).value:
            if from_excel_ordinal(worksheet.cell(row = i, column = 14).value) != from_excel_ordinal(worksheet.cell(row = i+1, column = 14).value):
                worksheet.insert_rows(i+1, 1)
                max_row += 1
                i -= 2
    
    return max_row
                
# https://stackoverflow.com/a/29387450
def from_excel_ordinal(ordinal, _epoch0=datetime(1899, 12, 31)):
    if int(ordinal) > 59:
        ordinal -= 1  # Excel leap year bug, 1900 is not a leap year!
    result = (_epoch0 + timedelta(days=ordinal)).replace(microsecond=0)
    return result.strftime('%d%m%Y')
    
def add_summary_lines(worksheet, parameter_column, title = '', field_column = 0):

    worksheet.insert_rows(2, 1)

    column = worksheet.max_column
    st_row = 1

    i = 3
    while worksheet.cell(row = i, column = parameter_column).value:
        if worksheet.cell(row = i, column = parameter_column).value:
            if worksheet.cell(row = i, column = parameter_column).value != worksheet.cell(row = i+1, column = parameter_column).value:
                worksheet.insert_rows(i+1, 2)
                worksheet.cell(row = i + 1, column = 1).value = worksheet.cell(row = i, column = 1).value
                worksheet.cell(row = i + 1, column = parameter_column).value = worksheet.cell(row = i, column = parameter_column).value
                if field_column > 0:
                    worksheet.cell(row = i + 1, column = field_column).number_format = '#,##0.00'
                    worksheet.cell(row = i + 1, column = field_column).value = get_column_sum(worksheet, column, st_row + 1, i)
                else:
                    worksheet.cell(row = i + 1, column = parameter_column + 1).number_format = '#,##0.00'
                    worksheet.cell(row = i + 1, column = parameter_column + 1).value = get_column_sum(worksheet, column, st_row + 1, i)
                if title != '':
                    worksheet.cell(row = i + 1, column = parameter_column).value = title
                st_row = i + 2
                i += 2
        i += 1
    
    return worksheet
    
def add_grand_total(worksheet):
    max_row = worksheet.max_row
    worksheet.cell(row = max_row + 2, column = 1).value = 'Grand Total'
    worksheet.cell(row = max_row + 2, column = 2).number_format = '#,##0.00'
    worksheet.cell(row = max_row + 2, column = 2).value = f"= SUM({get_total_sum(worksheet, 2, max_row)})"
    worksheet.column_dimensions['B'].width = 25
    worksheet.column_dimensions['H'].width = 25
    worksheet.column_dimensions['G'].width = 25
    return worksheet
    
def add_items(worksheet):
    worksheet.cell(row = 2, column = 1).value = worksheet.cell(row = 3, column = 1).value
    for i in range(0, worksheet.max_row):
        if worksheet.cell(row = i + 1, column = 2).value == None:
            worksheet.cell(row = i + 1, column = 1).value = worksheet.cell(row = i+2, column = 1).value
            worksheet.cell(row = i + 1, column = 2).value = worksheet.cell(row = i+2, column = 2).value
            
    return worksheet

def calculate_sums_buckets(worksheet, start_point, end_point):
    end_row = 0
    for i in range (start_point, end_point):
        start_row  = 3
        for j in range (3, worksheet.max_row):
            if worksheet.cell(row = j, column = 4).value == None:
                if worksheet.cell(row = j - 1, column = 4).value == None:
                    continue
                end_row = j - 1
                if start_row != 0:
                    worksheet.cell(row = j, column = i).value = get_column_sum(worksheet, column = i, st_row = start_row, max_row = end_row)
                start_row = j + 2                
    return worksheet