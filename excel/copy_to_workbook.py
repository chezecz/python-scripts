import openpyxl
import re

def copy_rows(source, target, pattern):
    row_count = 0
    max_row = 0
    # set up max row and column for data source file
    source_max_row = source.max_row
    source_max_column = source.max_column
    if (target.title == 'Data Invest'):
        column_offset = 0
    else:
        column_offset = 1
    # copy data from source to the template's copy
    for i in range(2, source_max_row +2):
        if re.match(pattern, str(source.cell(row = i-1, column = 4).value)):
            for j in range(1, source_max_column+1):
                row = i - row_count
                column = j + column_offset
                target.cell(row=row, column=column).value = source.cell(row = i-1, column = j).value
                max_row = row - 1
        else:
            row_count += 1
    return max_row
    
def copy_data(source, target, pattern):
    return copy_rows(source, target, pattern)
    
def insert_data(sheet, data, description):
    columns = [desc[0] for desc in description]
    sheet.append(columns)
    for row_object in data:
        row = [x for x in row_object]
        sheet.append(row)
    return sheet
    
