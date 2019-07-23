import openpyxl

def cleanup_jca(worksheet_reference, worksheet_cleaned):
    worksheet = delete_rows(worksheet_reference, worksheet_cleaned)
    return worksheet
    
def delete_rows(worksheet_reference, worksheet_cleaned):
    offset = 4
    worksheet_max_row = worksheet_reference.max_row
    worksheet_max_column = worksheet_reference.max_column
    for i in range(2, worksheet_max_row+1):
        if str(worksheet_reference.cell(row = i, column = 2).value) == "" or worksheet_reference.cell(row = i, column = 2).value == None:
            for j in range(2, worksheet_max_column+1):
                worksheet_cleaned.cell(i+offset, j).value = ""
                
    return worksheet_cleaned