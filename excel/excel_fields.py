import openpyxl

from db.connect_to_db import connect_to_db

def set_headers():
    results = {}
    result, headers = connect_to_db(query_headers(), 'guest')
    fields = [column[0] for column in headers]
    for row in result:
        results[row.product_code] = row.description 
    return results

    
def query_headers():
    return 'select product_code, description from prodcode'
    
def insert_section_description(worksheet):

    product_description = set_headers()
    max_row = worksheet.max_row
    
    for i in range(1, max_row + 1):
        if worksheet.cell(row = i, column = 1).value:
            if not worksheet.cell(row = i, column = 2).value:
                worksheet.cell(row = i, column = 2).value = product_description[worksheet.cell(row = i, column = 1).value]
                
    return worksheet