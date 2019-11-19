import openpyxl
import shutil
import warnings

# ignoring warnings created by openpyxl module
warnings.filterwarnings("ignore")

from utilities.get_dates import set_dates
from utilities.relative_to_absolute import update_path

from excel.convert_xls import convert_file
from excel.cleanup_jobcost import cleanup_file
from excel.copy_to_workbook import copy_data
from config.parse_config import parse_config
from excel.cleanup_jca import cleanup_jca
from excel.add_lines import add_blank_lines

def get_report():

    config = parse_config()
    config = config['FILEPATH']
    # setting up path to the template

    filepath_template = update_path(config["template"])

    # setting up path to the data source

    filepath_data = update_path(config["data"])
    filepath_data_output = update_path(config["data_output"])
    filepath_excel = update_path(config["excel"])

    dc = set_dates() # dc = date_content
    
    dc["filepath_output"] = update_path(dc["filepath_output"])
    
    # copy template
    # (to not lose the template)
    shutil.copy(filepath_template, dc["filepath_output"])

    # convert file
    filepath_data = convert_file(filepath_excel, filepath_data)

    # copy data file
    shutil.move(filepath_data, filepath_data_output)

    # creating workbook objects

    wb_template = openpyxl.load_workbook(filepath_template, read_only=True)
    wb_output = openpyxl.load_workbook(dc["filepath_output"])
    wb_data = openpyxl.load_workbook(filepath_data_output)

    # worksheet objects

    source = wb_data["Sheet1"]

    source = cleanup_file(source)

    for i in range (1, 8, 2):
        if wb_output.sheetnames[i+1] == "Job Analysis_COPAK" :
            pattern = "^[C]"
        elif wb_output.sheetnames[i+1] == "Job Analysis_PRINT":
            pattern = "^[JMDB]"
        else: 
            pattern = ".*"
        target=wb_output.worksheets[i]
        max_row = copy_data(source, target, pattern)
        if (i < 6):
            max_row = add_blank_lines(wb_output.worksheets[i], max_row)
            cleanup_jca(wb_output.worksheets[i], wb_output.worksheets[i+1])
            wb_output.worksheets[i+1].delete_rows(6+max_row, 500)
    # saving the output file

    wb_output.save(dc["filepath_output"])
    wb_data.save(filepath_data_output)

    return dc, source

if __name__ == '__main__':
    get_report()